"""
One-time setup script: Creates OpsBlueprint folder structure
and Excel tracker files on SharePoint.

Creates:
  OpsBlueprint/
    Leads/                              (lead attachments / documents)
    Proposals/                          (generated proposal PDFs)
    Invoices/Incoming/                  (vendor invoices to process)
    Invoices/Processed/                 (processed invoices)
    OpsBlueprint Leads Tracker.xlsx     (table: OBLeadsTable)
    OpsBlueprint Email Categories.xlsx  (table: OBCategories)

Run inside the python-helper container:
  docker compose exec n8n-python-helper python /scripts/setup-opsblueprint-sharepoint.py

Or mount and run:
  docker compose run --rm -v ./n8n-agent/scripts:/scripts n8n-python-helper python /scripts/setup-opsblueprint-sharepoint.py
"""

import io
import json
import os
import sys
import time
import urllib.request
import urllib.parse

# -- Configuration ------------------------------------------------------------

TENANT_ID = os.environ.get("MS_GRAPH_TENANT_ID", "")
CLIENT_ID = os.environ.get("MS_GRAPH_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("MS_GRAPH_CLIENT_SECRET", "")
DRIVE_ID = os.environ.get("SHAREPOINT_DRIVE_ID", "")

if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, DRIVE_ID]):
    print("ERROR: Missing environment variables. Need:")
    print("  MS_GRAPH_TENANT_ID, MS_GRAPH_CLIENT_ID, MS_GRAPH_CLIENT_SECRET, SHAREPOINT_DRIVE_ID")
    print()
    print("Run this inside the n8n-python-helper container which has these set,")
    print("or pass them via: docker compose run --rm -e MS_GRAPH_TENANT_ID=... n8n-python-helper ...")
    sys.exit(1)

BASE_FOLDER = "OpsBlueprint"


# -- Graph API helpers ---------------------------------------------------------

def get_graph_token() -> str:
    """Acquire MS Graph token via client_credentials OAuth."""
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }).encode()

    req = urllib.request.Request(url, data=data, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")

    with urllib.request.urlopen(req) as resp:
        body = json.loads(resp.read())
    return body["access_token"]


def graph_get(token: str, path: str) -> dict:
    """GET request to MS Graph API."""
    url = f"https://graph.microsoft.com/v1.0{path}"
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Bearer {token}")
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def graph_post(token: str, path: str, payload: dict) -> dict:
    """POST request to MS Graph API."""
    url = f"https://graph.microsoft.com/v1.0{path}"
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, method="POST")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def graph_patch(token: str, path: str, payload: dict) -> dict:
    """PATCH request to MS Graph API."""
    url = f"https://graph.microsoft.com/v1.0{path}"
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, method="PATCH")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def graph_put_bytes(token: str, path: str, content: bytes, content_type: str) -> dict:
    """PUT binary content to MS Graph API."""
    url = f"https://graph.microsoft.com/v1.0{path}"
    req = urllib.request.Request(url, data=content, method="PUT")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", content_type)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


# -- Folder creation -----------------------------------------------------------

def get_or_create_folder(token: str, parent_path: str, folder_name: str) -> str:
    """Get existing folder or create it. Returns item ID."""
    encoded_path = urllib.parse.quote(f"{parent_path}/{folder_name}")
    try:
        result = graph_get(token, f"/drives/{DRIVE_ID}/root:/{encoded_path}")
        print(f"    Found existing folder: {folder_name} (id: {result['id']})")
        return result["id"]
    except urllib.error.HTTPError:
        pass

    encoded_parent = urllib.parse.quote(parent_path)
    result = graph_post(token, f"/drives/{DRIVE_ID}/root:/{encoded_parent}:/children", {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "fail"
    })
    print(f"    Created folder: {folder_name} (id: {result['id']})")
    return result["id"]


# -- Excel file creation -------------------------------------------------------

def create_leads_tracker_xlsx() -> bytes:
    """Create OpsBlueprint Leads Tracker Excel with OBLeadsTable."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "LeadID", "FullName", "Email", "Company", "Industry",
        "ProblemDescription", "ServiceTier", "Source",
        "QualificationScore", "QualificationNotes",
        "SuggestedTier", "EstimatedValue", "Urgency",
        "FollowUpRecommendation", "Status", "ProposalSent",
        "CreatedAt"
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # One sample row to ensure table creation works
    sample = [
        "OB-SAMPLE-001", "Jane Doe", "jane@example.com", "Sample Corp",
        "E-commerce", "Manual order processing across 3 platforms",
        "Professional", "opsblueprint.fernando-vargas.com",
        "8", "Strong lead: multiple platforms, clear pain point",
        "Professional", "$8,000", "medium",
        "Schedule discovery call to map order workflow",
        "New", "", "2026-03-18T00:00:00Z"
    ]
    for col, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_email_categories_xlsx() -> bytes:
    """Create OpsBlueprint Email Categories Excel with OBCategories table."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"

    headers = ["Category", "Description", "FolderName", "Priority"]

    categories = [
        [
            "New Leads",
            "Inquiries about automation consulting services, RFPs, pricing "
            "requests, discovery call requests from prospective clients",
            "New Leads",
            "High"
        ],
        [
            "Project Updates",
            "Communications about ongoing automation projects, status updates, "
            "milestone completions, change requests, scope discussions",
            "Project Updates",
            "High"
        ],
        [
            "Support",
            "Issues with deployed automations, error reports, bug fixes, "
            "performance problems, workflow failures, urgent requests",
            "Support",
            "High"
        ],
        [
            "Billing & Contracts",
            "Invoice inquiries, payment confirmations, contract renewals, "
            "managed service billing, scope change approvals",
            "Billing",
            "Medium"
        ],
        [
            "Partner & Vendor",
            "Communications from technology partners, platform vendors, "
            "API providers, referral partners, integration marketplace",
            "Partners",
            "Medium"
        ],
        [
            "Feedback & Reviews",
            "Client feedback, testimonials, case study requests, satisfaction "
            "surveys, referral requests, post-project reviews",
            "Feedback",
            "Low"
        ],
        [
            "Uncategorized",
            "Default for emails that don't clearly fit other categories, "
            "newsletters, spam, general correspondence",
            "Uncategorized",
            "Low"
        ],
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(categories, start=2):
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_and_create_table(
    token: str,
    file_name: str,
    xlsx_bytes: bytes,
    sheet_name: str,
    table_name: str,
    num_cols: int,
    num_data_rows: int,
) -> str:
    """Upload Excel file to SharePoint and create a named table. Returns file item ID."""
    encoded_path = urllib.parse.quote(f"{BASE_FOLDER}/{file_name}")
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Upload
    result = graph_put_bytes(
        token,
        f"/drives/{DRIVE_ID}/root:/{encoded_path}:/content",
        xlsx_bytes,
        content_type,
    )
    item_id = result["id"]
    print(f"    Uploaded '{file_name}' -> id: {item_id}")

    # Wait for Excel service to index the file
    time.sleep(3)

    # Create named table from data range
    total_rows = 1 + num_data_rows  # header + data
    last_col = chr(ord("A") + num_cols - 1) if num_cols <= 26 else "Q"
    address = f"{sheet_name}!A1:{last_col}{total_rows}"

    table_result = graph_post(
        token,
        f"/drives/{DRIVE_ID}/items/{item_id}/workbook/tables/add",
        {"address": address, "hasHeaders": True},
    )
    auto_name = table_result.get("name", "")
    print(f"    Created table '{auto_name}' from range {address}")

    # Rename table if needed
    if auto_name != table_name:
        graph_patch(
            token,
            f"/drives/{DRIVE_ID}/items/{item_id}/workbook/tables/{auto_name}",
            {"name": table_name},
        )
        print(f"    Renamed table to '{table_name}'")

    return item_id


# -- Main ----------------------------------------------------------------------

def main():
    print("=" * 60)
    print("OpsBlueprint - SharePoint Setup")
    print("=" * 60)
    print()

    print("[1/7] Acquiring Graph token...")
    token = get_graph_token()
    print("  Token acquired")

    print()
    print("[2/7] Ensuring base folder exists...")
    try:
        encoded = urllib.parse.quote(BASE_FOLDER)
        base = graph_get(token, f"/drives/{DRIVE_ID}/root:/{encoded}")
        base_id = base["id"]
        print(f"  Found '{BASE_FOLDER}' (id: {base_id})")
    except urllib.error.HTTPError:
        print(f"  '{BASE_FOLDER}' not found -- creating it...")
        result = graph_post(token, f"/drives/{DRIVE_ID}/root/children", {
            "name": BASE_FOLDER,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "fail"
        })
        base_id = result["id"]
        print(f"  Created '{BASE_FOLDER}' (id: {base_id})")

    print()
    print("[3/7] Creating sub-folders...")
    leads_folder_id = get_or_create_folder(token, BASE_FOLDER, "Leads")
    proposals_folder_id = get_or_create_folder(token, BASE_FOLDER, "Proposals")
    get_or_create_folder(token, BASE_FOLDER, "Invoices")
    invoices_incoming_id = get_or_create_folder(
        token, f"{BASE_FOLDER}/Invoices", "Incoming"
    )
    invoices_processed_id = get_or_create_folder(
        token, f"{BASE_FOLDER}/Invoices", "Processed"
    )

    print()
    print("[4/7] Generating Leads Tracker Excel...")
    leads_xlsx = create_leads_tracker_xlsx()
    print(f"  Generated {len(leads_xlsx)} bytes")

    print()
    print("[5/7] Uploading Leads Tracker to SharePoint...")
    leads_file_id = upload_and_create_table(
        token,
        "OpsBlueprint Leads Tracker.xlsx",
        leads_xlsx,
        sheet_name="Leads",
        table_name="OBLeadsTable",
        num_cols=17,
        num_data_rows=1,
    )

    print()
    print("[6/7] Generating Email Categories Excel...")
    categories_xlsx = create_email_categories_xlsx()
    print(f"  Generated {len(categories_xlsx)} bytes")

    print()
    print("[7/7] Uploading Email Categories to SharePoint...")
    categories_file_id = upload_and_create_table(
        token,
        "OpsBlueprint Email Categories.xlsx",
        categories_xlsx,
        sheet_name="Categories",
        table_name="OBCategories",
        num_cols=4,
        num_data_rows=7,
    )

    # -- Output ----------------------------------------------------------------

    print()
    print("=" * 60)
    print("SUCCESS! Add these to your deploy/.env file:")
    print("=" * 60)
    print()
    print(f"OB_SHAREPOINT_LEADS_FOLDER_ID={leads_folder_id}")
    print(f"OB_SHAREPOINT_PROPOSALS_FOLDER_ID={proposals_folder_id}")
    print(f"OB_SHAREPOINT_INVOICES_INCOMING_FOLDER_ID={invoices_incoming_id}")
    print(f"OB_SHAREPOINT_INVOICES_PROCESSED_FOLDER_ID={invoices_processed_id}")
    print(f"OB_LEADS_TRACKER_FILE_ID={leads_file_id}")
    print(f"OB_EMAIL_CATEGORIES_FILE_ID={categories_file_id}")
    print(f"OB_EMAIL_CATEGORIES_TABLE_NAME=OBCategories")
    print()
    print("Then restart n8n: docker compose restart n8n")
    print("=" * 60)


if __name__ == "__main__":
    main()
